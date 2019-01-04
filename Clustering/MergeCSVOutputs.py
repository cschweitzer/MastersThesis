# -*- coding: utf-8 -*-
"""
Created on Fri Jan  4 13:39:11 2019

@author: schwe
"""


import pandas as pd
import os
import csv
import glob
#%%
clusterfiles = glob.glob(os.path.join('..\Outputs\Clustering', "*.csv"))

writer = pd.ExcelWriter('../Outputs/Clustering/combined_cluster_outputs_samesheet.xlsx')

df = pd.DataFrame()

for f in clusterfiles:
    temp = pd.read_csv(f)
    clustertypeval = os.path.basename(f).replace(".csv","").replace("summary_","")

    temp.insert(0, 'clustertype',
              value = clustertypeval)
    df = df.append(temp, sort = False)

df.to_excel(writer,sheet_name = 'summaries',
                index = False)
writer.save()
writer.close()

#%%

writer = pd.ExcelWriter('../Outputs/Clustering/combined_cluster_outputs.xlsx')

import pandas as pd
from openpyxl import load_workbook

book = load_workbook('test.xlsx')
writer = pd.ExcelWriter('test.xlsx')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

df.to_excel(writer, sheet_name='tab_name', other_params)

writer.save()


#%% add to existing

from openpyxl import load_workbook

path = r"../Outputs/Clustering/combined_cluster_outputs.xlsx"

book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book

df = pd.read_csv(f)

df.to_excel(writer, sheet_name = '')

writer.save()
writer.close()